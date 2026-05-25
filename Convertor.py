import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
import os

def get_safe_unique_title(title, existing_titles):
    """
    生成一个符合 Excel 规范且唯一的 Sheet 名称。
    1. 替换非法字符 \ / * [ ] : ?
    2. 截断至 31 字符
    3. 处理重名，通过添加 _1, _2 等后缀确保唯一性
    """
    safe_title = title
    for char in r'\/*[]:?':
        safe_title = safe_title.replace(char, ' ')

    safe_title = safe_title[:31].strip() or "Untitled Book"

    final_title = safe_title
    counter = 1
    while final_title in existing_titles:
        base_name = safe_title[:25]
        suffix = f"_{counter}"
        final_title = (base_name + suffix)[:31]
        counter += 1

    existing_titles.add(final_title)
    return final_title

def group_data_by_book(file_path):
    """
    读取 CSV 文件并按书名将记录分组。
    """
    books_data = defaultdict(list)
    if not os.path.exists(file_path):
        print(f"错误: 找不到输入文件 {file_path}")
        return None

    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                book = row.get('书名', '未知书名')
                books_data[book].append(row)
    except Exception as e:
        print(f"读取 CSV 时出错: {e}")
        return None

    return books_data

def sort_entries(entries):
    """
    按 '位置' 栏目的数值从小到大进行升序排序。
    """
    def get_loc_val(entry):
        loc_str = entry.get('位置', '0')
        try:
            clean_loc = ''.join(filter(lambda x: x.isdigit() or x == '-', loc_str))
            if '-' in clean_loc:
                return int(clean_loc.split('-')[0])
            return int(clean_loc) if clean_loc else 0
        except ValueError:
            return 0

    return sorted(entries, key=get_loc_val)

def generate_filename():
    """
    生成文件名：Clippings-YYYYMMDDHHMMSS.xlsx
    """
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"Clippings-{timestamp}.xlsx"

def setup_sheet_layout(ws):
    """
    设置表头、字体样式和列宽。
    """
    headers = ["书名", "位置", "页码", "摘录", "笔记", "日期时间"]
    widths = [30, 15, 10, 100, 50, 20]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

def write_all_data_to_sheet(ws, entries):
    """
    将所有排序后的数据写入同一个工作表，并设置字体大小、自动换行和动态行高。
    """
    sorted_entries = sort_all_data(entries) 

    for row_idx, entry in enumerate(sorted_entries, 2):
        row_data = [
            entry.get('书名', ''),
            entry.get('位置', ''),
            entry.get('页码', ''),
            entry.get('摘录内容', ''),
            entry.get('笔记内容', ''),
            entry.get('日期时间', '')
        ]

        for col_idx, value in enumerate(row_data, 1): #单元格对齐格式修改
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=12)
            if col_idx in [4, 5]:
                cell.alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
            else:
                cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        excerpt_text = entry.get('摘录内容', '')
        if excerpt_text:
            line_count = (len(excerpt_text) // 60) + 1
            calculated_height = 16 * line_count + 5
            ws.row_dimensions[row_idx].height = min(calculated_height, 300)
        else:
            ws.row_dimensions[row_idx].height = 16

def sort_all_data(entries):
    """
    二级排序：先按书名排序，再在同一本书内按位置数值从小到大排序。
    """
    def get_loc_val(entry):
        loc_str = entry.get('位置', '0')
        try:
            clean_loc = ''.join(filter(lambda x: x.isdigit() or x == '-', loc_str))
            if '-' in clean_loc:
                return int(clean_loc.split('-')[0])
            return int(clean_loc) if clean_loc else 0
        except ValueError:
            return 0

    return sorted(entries, key=lambda x: (x.get('书名', ''), get_loc_val(x)))

def finalize_sheet(ws, total_rows):
    """
    冻结第一行并仅在 A 列开启自动筛选。
    """
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:A{total_rows + 1}"

def main():
    input_file = "My Clippings.csv"
    print(f"正在读取 {input_file}...")

    all_data = group_data_by_book(input_file) 
    all_entries = []
    if not os.path.exists(input_file):
        print(f"错误: 找不到输入文件 {input_file}")
        return False

    try:
        with open(input_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                all_entries.append(row)
    except Exception as e:
        print(f"读取 CSV 时出错: {e}")
        return False

    if not all_entries:
        print("无法读取数据，程序退出。")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    today_date = datetime.now().strftime("%Y%m%d")
    ws.title = today_date

    print(f"正在生成工作表 {today_date} 并设置格式...")

    setup_sheet_layout(ws)
    write_all_data_to_sheet(ws, all_entries)
    finalize_sheet(ws, len(all_entries))

    output_file = generate_filename()
    try:
        wb.save(output_file)
        print(f"\n成功！文件已保存为: {output_file}")
        print(f"总计导出 {len(all_entries)} 条记录到单一工作表。")
        return True
    except Exception as e:
        print(f"保存文件时出错: {e}")
        return False

if __name__ == "__main__":
    main()
